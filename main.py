import cv2
import os
import pickle
import face_recognition
import numpy as np
import cvzone
import firebase_admin
from firebase_admin import db
from firebase_admin import credentials
from firebase_admin import storage
import datetime
import openpyxl
from openpyxl import load_workbook

#wb = openpyxl.Workbook()
wb = load_workbook('farts.xlsx')
sheet = wb.active
row = 1
col = 1

cred = credentials.Certificate("serviceAccountKey.json")

firebase_admin.initialize_app(cred, {
    'databaseURL':"https://facerecognitionrealtime-e783a-default-rtdb.firebaseio.com/",
    'storageBucket':"facerecognitionrealtime-e783a.appspot.com"
}) 

bucket = storage.bucket()

cap = cv2.VideoCapture(0)
cap.set(3, 640)
cap.set(4, 480)

imgBackground = cv2.imread("Resources/background.png")

#Importing Modes images into a list 
folderModePath = "Resources/Modes"
modePathList = os.listdir(folderModePath)
#List of images of mode
imgModeList = []
for path in modePathList:
    imgModeList.append(cv2.imread(os.path.join(folderModePath,path)))

#print(len(imgModeList))
#Load the Encoding File
print("Loading Encoded File...")
file = open("EncodeFile.p","rb")
encodeListKnownWithIds = pickle.load(file)
file.close()

#Mapping data loaded as desired
encodeListKnown, studentIds = encodeListKnownWithIds
print("Encoded File Loaded")
#print(studentIds)

modeType = 0
counter = 0
id = -1
imgStudent = []

while(True):
    success, img = cap.read()
    
    #Squeezing image to avoid overhead computations
    imgS = cv2.resize(img,(0,0),None,0.25,0.25) #Scaling Down
    imgS = cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
    
    #We need two things as current frame and encoding in the current frame
    faceCurrframe = face_recognition.face_locations(imgS);
    encodeCurrFrame = face_recognition.face_encodings(imgS,faceCurrframe)
    
    #Overlaying background.png on webcam
    imgBackground[162:162+480, 55:55+640] = img
    #Meshing mode images with the background
    imgBackground[44:44+633, 808:808+414] = imgModeList[modeType]
    
    #Loop through above generated encodings and comparing them
    if faceCurrframe:
        for encodeFace, faceLoc in zip(encodeCurrFrame, faceCurrframe):
            matches = face_recognition.compare_faces(encodeListKnown,encodeFace) #Direct Proportion
            faceDist = face_recognition.face_distance(encodeListKnown, encodeFace)#Inverse Porpotion
            
            matchIndex = np.argmin(faceDist)
            
            if matches[matchIndex]:
                # print("Known Face Detected : ",studentIds[matchIndex])
                y1, x1, y2, x2 = faceLoc
                y1, x1, y2, x2 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                bbox = 55 + x1, 162 + y1, x2 - x1, y2 - y1
                imgBackground = cvzone.cornerRect(imgBackground, bbox, rt=0)
                id = studentIds[matchIndex]
                if counter==0:
                    cvzone.putTextRect(imgBackground,"Loading",(275,400))
                    cv2.imshow("Face Attendance", imgBackground)
                    cv2.waitKey(1)
                    counter = 1
                    modeType = 1

        if counter!=0:
            if counter==1:
                
                #Downloading Data from Storage
                studentInfo = db.reference(f'Students/{id}').get()
                print(studentInfo)
                
                #Get Images from Storage
                blob = bucket.get_blob(f'Images/{id}.jpg')
                array = np.frombuffer(blob.download_as_string(), np.uint8)
                imgStudent = cv2.imdecode(array, cv2.COLOR_BGR2RGB)
                datetime_object = datetime.datetime.strptime(studentInfo['last_attendance_time'], "%Y-%m-%d %H:%M:%S")
                
                secondsElapsed = (datetime.datetime.now() - datetime_object).total_seconds()
                print(secondsElapsed)
                #Update data of Attendance
                if secondsElapsed>45:
                    flag = 0
                    while(flag==0):
                        #print("Already : ",sheet.cell(row,col).value)
                        if(sheet.cell(row,col).value==None):
                            sheet.cell(row,col).value = studentInfo['Name']
                            col += 1
                            sheet.cell(row,col).value = studentInfo['last_attendance_time']
                            flag = 1
                            break
                        else:
                            row += 1
                    flag = 0
                    row = 1
                    col = 1
                    ref = db.reference(f'Students/{id}')
                    studentInfo['Total Attendance'] += 1
                    ref.child('Total Attendance').set(studentInfo['Total Attendance'])
                    ref.child('last_attendance_time').set(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    modeType = 3
                    counter = 0
                    imgBackground[44:44+633, 808:808+414] = imgModeList[modeType]
                    
            if modeType!=3:
                if 8<counter<=28:
                    modeType = 2 
                    imgBackground[44:44+633, 808:808+414] = imgModeList[modeType]   
            
                if counter<=18:   
                    cv2.putText(imgBackground,str(studentInfo['Total Attendance']),(861,125),cv2.FONT_HERSHEY_COMPLEX,1,(255,255,255),1)
                    cv2.putText(imgBackground,str(studentInfo['Department']),(1006,550),cv2.FONT_HERSHEY_COMPLEX,0.5,(255,255,255),1)
                    cv2.putText(imgBackground,str(id),(1006,493),cv2.FONT_HERSHEY_COMPLEX,1,(255,255,255),1)
                    cv2.putText(imgBackground,str(studentInfo['Div']),(910,625),cv2.FONT_HERSHEY_COMPLEX,0.6,(100,100,100),1)
                    cv2.putText(imgBackground,str(studentInfo['Class']),(1025,625),cv2.FONT_HERSHEY_COMPLEX,0.6,(100,100,100),1)
                    (w,h) , _ = cv2.getTextSize(studentInfo['Name'],cv2.FONT_HERSHEY_COMPLEX,1,1)
                    offset = (414-w)//2
                    cv2.putText(imgBackground,str(studentInfo['Name']),(808+offset,445),cv2.FONT_HERSHEY_COMPLEX,1,(50,50,50),1)
                    
                    imgBackground[175:175+216, 909:909+216] = imgStudent
            counter += 1
            if counter>=15:
                counter = 0
                modeType = 0
                studentInfo = []
                imgStudent = []
                imgBackground[44:44+633, 808:808+414] = imgModeList[modeType]
    else:
        modeType = 0
        counter = 0
            
    #cv2.imshow("Webcam",img)
    wb.save("farts.xlsx")
    #print("Done")
    cv2.imshow("Face Attendance", imgBackground)
    cv2.waitKey(1)